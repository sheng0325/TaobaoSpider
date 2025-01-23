

import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from pyquery import PyQuery as pq

import openpyxl
from openpyxl import Workbook, load_workbook

# ===== 全局变量 =====
count = 2  # Excel写入行数的计数，假设第1行是表头，从第2行开始写
wb = None  # 全局工作簿对象
ws = None  # 全局工作表对象

def init_excel(file_name, sheet_name="Sheet1"):
    """
    初始化Excel。如果文件不存在则创建，并写入表头；
    如果文件存在则直接加载（并假设已包含表头）。
    返回 (wb, ws) 对象
    """
    if not os.path.exists(file_name):
        print(f"Excel 文件 [{file_name}] 不存在，将新建并写入表头...")
        wb_new = Workbook()
        ws_new = wb_new.active
        ws_new.title = sheet_name

        # 写入表头，按你需要的字段自定义
        ws_new.cell(row=1, column=1, value='Page')
        ws_new.cell(row=1, column=2, value='Num')
        ws_new.cell(row=1, column=3, value='title')
        ws_new.cell(row=1, column=4, value='price')
        ws_new.cell(row=1, column=5, value='deal')
        ws_new.cell(row=1, column=6, value='location')
        ws_new.cell(row=1, column=7, value='shop')
        ws_new.cell(row=1, column=8, value='isPostFree')
        ws_new.cell(row=1, column=9, value='url')
        ws_new.cell(row=1, column=10, value='shop_url')
        ws_new.cell(row=1, column=11, value='img_url')
        ws_new.cell(row=1, column=12, value='style1')
        ws_new.cell(row=1, column=13, value='style2')
        ws_new.cell(row=1, column=14, value='style3')

        wb_new.save(file_name)
        return wb_new, ws_new
    else:
        print(f"Excel 文件 [{file_name}] 已存在，将直接加载...")
        wb_exist = load_workbook(file_name)
        if sheet_name in wb_exist.sheetnames:
            ws_exist = wb_exist[sheet_name]
        else:
            ws_exist = wb_exist.create_sheet(sheet_name)
            # 如果新 sheet，需要写表头
            ws_exist.cell(row=1, column=1, value='Page')
            ws_exist.cell(row=1, column=2, value='Num')
            ws_exist.cell(row=1, column=3, value='title')
            ws_exist.cell(row=1, column=4, value='price')
            ws_exist.cell(row=1, column=5, value='deal')
            ws_exist.cell(row=1, column=6, value='location')
            ws_exist.cell(row=1, column=7, value='shop')
            ws_exist.cell(row=1, column=8, value='isPostFree')
            ws_exist.cell(row=1, column=9, value='url')
            ws_exist.cell(row=1, column=10, value='shop_url')
            ws_exist.cell(row=1, column=11, value='img_url')
            ws_exist.cell(row=1, column=12, value='style1')
            ws_exist.cell(row=1, column=13, value='style2')
            ws_exist.cell(row=1, column=14, value='style3')

        return wb_exist, ws_exist

def init_driver(driver_path, headless=False):
    """
    初始化 Selenium WebDriver
    """
    chrome_options = Options()
    if headless:
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')
    # 根据需要添加更多选项
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.maximize_window()
    return driver

def login_taobao(driver, username, password):
    """
    简易示例：登录淘宝（可能需要手动滑块、验证码等）
    """
    print("1) 打开淘宝登录页")
    driver.get("https://login.taobao.com/")
    time.sleep(2)

    try:
        print("2) 切换到密码登录页面（若有）")
        pwd_login_tab = driver.find_element(By.XPATH, '//a[@class="forget-pwd J_Quick2Static"]')
        pwd_login_tab.click()
        time.sleep(1)
    except NoSuchElementException:
        print("   - 没有找到切换到密码登录的按钮，可能已经是密码登录界面。")

    print("3) 输入用户名和密码")
    user_input = driver.find_element(By.ID, "fm-login-id")
    user_input.clear()
    user_input.send_keys(username)
    time.sleep(0.5)

    pwd_input = driver.find_element(By.ID, "fm-login-password")
    pwd_input.clear()
    pwd_input.send_keys(password)
    time.sleep(0.5)

    print("4) 点击登录按钮")
    login_btn = driver.find_element(By.XPATH, '//button[@type="submit"]')
    login_btn.click()

    print("如果出现验证码或滑块，需要你手动处理；处理完后按回车继续")
    input("   按回车继续...")

    # 等待一下，给页面跳转时间
    time.sleep(3)
    print("   - 登录流程结束")

def get_goods(page, driver, wb, ws):
    """
    根据你给出的示例编写的函数：
    1. 等待用户手动确认“页面加载完毕”，输入数字“1”开始爬取
    2. 使用 PyQuery 解析当前页面 HTML
    3. 提取商品信息并写入 Excel
    """
    global count  # 引用全局计数

    # 手动确认页面加载完毕
    user_in = input('确认界面加载完毕，输入数字“1”开始爬取-->')
    # Python里 input() 返回字符串，需要判断 == '1' 而不是 1
    if user_in.strip() == '1':
        pass
    else:
        print("   - 非“1”，跳过爬取。")
        return

    # 获取当前页面HTML
    html = driver.page_source
    doc = pq(html)

    # 提取所有商品外层选择器（根据你给的示例）
    items = doc('div.content--CUnfXXxv > div > div').items()
    for item in items:
        # 标题
        title = item.find('.title--qJ7Xg_90 span').text()
        # 价格
        price_int = item.find('.priceInt--yqqZMJ5a').text()
        price_float = item.find('.priceFloat--XpixvyQ1').text()
        if price_int and price_float:
            price = float(f"{price_int}{price_float}")
        else:
            price = 0.0
        # 交易量
        deal = item.find('.realSales--XZJiepmt').text()
        # 所在地
        location = item.find('.procity--wlcT2xH9 span').text()
        # 店名
        shop = item.find('.shopNameText--DmtlsDKm').text()
        # 包邮？
        postText = item.find('.subIconWrapper--Vl8zAdQn').text()
        postText = "包邮" if "包邮" in postText else "/"
        # 商品URL
        t_url_el = item.find('.doubleCardWrapperAdapt--mEcC7olq')
        t_url = t_url_el.attr('href') if t_url_el else ''
        # 店铺URL
        shop_url_el = item.find('.TextAndPic--grkZAtsC a')
        shop_url = shop_url_el.attr('href') if shop_url_el else ''
        # 图片URL
        img_el = item.find('.mainPicAdaptWrapper--V_ayd2hD img')
        img_url = img_el.attr('src') if img_el else ''

        # 风格
        style_list = item('div.abstractWrapper--whLX5va5 > div').items()
        style = []
        for s in style_list:
            s_span = s('div.descBox--RunOO4S3 > span').text()
            if s_span:
                style.append(s_span)

        # 构建字典（可选，调试用）
        product = {
            'Page':       page,
            'Num':        count - 1,
            'title':      title,
            'price':      price,
            'deal':       deal,
            'location':   location,
            'shop':       shop,
            'isPostFree': postText,
            'url':        t_url,
            'shop_url':   shop_url,
            'img_url':    img_url,
            'style':      style
        }
        print(product)

        # 写入Excel
        ws.cell(row=count, column=1, value=page)           # 页码
        ws.cell(row=count, column=2, value=(count - 1))    # 序号
        ws.cell(row=count, column=3, value=title)          # 标题
        ws.cell(row=count, column=4, value=price)          # 价格
        ws.cell(row=count, column=5, value=deal)           # 付款人数
        ws.cell(row=count, column=6, value=location)       # 地理位置
        ws.cell(row=count, column=7, value=shop)           # 店铺名称
        ws.cell(row=count, column=8, value=postText)       # 是否包邮
        ws.cell(row=count, column=9, value=t_url)          # 商品链接
        ws.cell(row=count, column=10, value=shop_url)      # 商铺链接
        ws.cell(row=count, column=11, value=img_url)       # 图片链接
        # 写入风格（最多 3 个）
        for i in range(min(3, len(style))):
            ws.cell(row=count, column=12 + i, value=style[i])

        count += 1  # 行数+1

def task1():
    """
    任务1：示例
    1. 初始化Excel
    2. 初始化driver
    3. 登录淘宝
    4. 搜索关键词
    5. 每页调用get_goods(page)
    6. 结束后保存Excel并退出
    """
    global wb, ws, count

    # ========== 你可以根据自己实际情况修改这些参数 ==========
    EXCEL_FILE = "goods_data.xlsx"          # Excel 文件名
    SHEET_NAME = "Task1"                    # Sheet名
    DRIVER_PATH = "./chromedriver"          # ChromeDriver 路径
    USERNAME = "test_user"                  # 淘宝账号
    PASSWORD = "test_pass"                  # 淘宝密码
    SEARCH_KEYWORD = "洗衣液"               # 要搜索的关键词
    MAX_PAGES = 2                           # 爬多少页
    HEADLESS = False                        # 是否无头
    # ========================================================

    # 1) 初始化 Excel
    wb, ws = init_excel(EXCEL_FILE, SHEET_NAME)
    # 2) 初始化 driver
    driver = init_driver(DRIVER_PATH, HEADLESS)
    # 3) 登录淘宝
    login_taobao(driver, USERNAME, PASSWORD)

    print(f"现在搜索商品：{SEARCH_KEYWORD}")
    driver.get("https://s.taobao.com/")
    time.sleep(2)

    # 输入关键词后搜索
    search_input = driver.find_element(By.ID, "q")
    search_input.clear()
    search_input.send_keys(SEARCH_KEYWORD)
    search_input.send_keys(Keys.ENTER)
    time.sleep(3)

    # 4) 爬取多页
    for page in range(1, MAX_PAGES + 1):
        print(f"=== 处理第 {page} / {MAX_PAGES} 页 ===")
        # 每页先向下滚动，触发懒加载
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)

        # 调用 get_goods(page)
        get_goods(page, driver, wb, ws)

        # 点击下一页（如果不是最后一页）
        if page < MAX_PAGES:
            try:
                next_btn = driver.find_element(By.CSS_SELECTOR, 'span.next-btn-helper')
                next_btn.click()
                time.sleep(3)
            except NoSuchElementException:
                print("   - 未找到下一页按钮，提前结束。")
                break

    # 5) 保存Excel
    wb.save(EXCEL_FILE)
    print(f"Excel 已保存到 {EXCEL_FILE}")
    # 关闭浏览器
    driver.quit()
    print("任务1已完成并退出程序。")

if __name__ == "__main__":
    """
    如果你只想运行这一个任务，运行 python task1.py 即可。
    任务执行完毕后自动结束程序。
    """
    task1()
