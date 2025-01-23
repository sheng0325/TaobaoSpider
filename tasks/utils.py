# tasks/utils.py

import os
import time
import json
import yaml
import openpyxl
from pyquery import PyQuery as pq
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException

def load_config(config_path: str = "config.yaml") -> dict:
    """
    加载 config.yaml
    """
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

def load_accounts(accounts_path: str = "accounts.json") -> list:
    """
    加载 accounts.json
    返回一个 list，里面每个元素是 {username, password, type}
    """
    with open(accounts_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def init_webdriver(driver_path: str, headless: bool = False):
    """
    初始化并返回Chrome WebDriver
    """
    chrome_options = Options()
    if headless:
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')
    # 视实际需求添加更多的 Options

    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.maximize_window()
    return driver

def login_taobao(driver, username: str, password: str, manual_verify_wait: int = 60):
    """
    登录淘宝示例：可能会遇到验证码、滑块、短信、扫码等复杂验证
    这里只做基本示例，若出现验证，则手动处理后回车继续
    """
    print("正在执行：进入淘宝登录页")
    driver.get("https://login.taobao.com/")
    time.sleep(2)

    # 如果默认进入的是扫码界面，尝试切换到 密码登录
    try:
        pwd_login_tab = driver.find_element(By.XPATH, '//a[@class="forget-pwd J_Quick2Static"]')
        pwd_login_tab.click()
        time.sleep(1)
    except:
        # 如果找不到，则可能已经是密码登录界面
        pass

    print("正在执行：输入用户名")
    user_input = driver.find_element(By.ID, "fm-login-id")
    user_input.clear()
    user_input.send_keys(username)
    time.sleep(1)

    print("正在执行：输入密码")
    pwd_input = driver.find_element(By.ID, "fm-login-password")
    pwd_input.clear()
    pwd_input.send_keys(password)
    time.sleep(1)

    print("正在执行：点击登录按钮")
    login_button = driver.find_element(By.XPATH, '//*[@id="login-form"]/div[6]/button')
    login_button.click()

    print(f"如果出现验证码或滑块，请手动处理，处理完后按回车继续（等待最多 {manual_verify_wait} 秒）...")
    input("按回车键继续...")
    # 简单等待，给页面跳转足够时间
    time.sleep(3)

def scroll_to_bottom(driver, times=3, sleep_interval=1):
    """
    向下滚动页面多次，用于触发懒加载
    """
    for i in range(times):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(sleep_interval)

def parse_item_info(item):
    item = pq(item)  # 将 item 转换为 PyQuery 对象

    # 定位商品标题
    title = item.find('.title--qJ7Xg_90 span').text()

    # 定位价格
    price_int = item.find('.priceInt--yqqZMJ5a').text()
    price_float = item.find('.priceFloat--XpixvyQ1').text()
    price = float(f"{price_int}{price_float}") if price_int and price_float else 0.0

    # 定位交易量
    deal = item.find('.realSales--XZJiepmt').text()

    # 定位所在地信息
    location = item.find('.procity--wlcT2xH9 span').text()

    # 定位店名
    shop = item.find('.shopNameText--DmtlsDKm').text()

    # 定位包邮的位置
    postText_raw = item.find('.subIconWrapper--Vl8zAdQn').text()
    postText = "包邮" if "包邮" in postText_raw else "/"

    # 定位商品url
    t_url = item.find('.doubleCardWrapperAdapt--mEcC7olq').attr('href')

    # 定位店名url
    shop_url = item.find('.TextAndPic--grkZAtsC a').attr('href')

    # 定位商品图片url
    img_url = item.find('.mainPicAdaptWrapper--V_ayd2hD img').attr('src')

    return {
        "title": title,
        "price": price,
        "deal": deal,
        "location": location,
        "shop": shop,
        "postText": postText,
        "t_url": t_url,
        "shop_url": shop_url,
        "img_url": img_url
    }

def save_to_excel(excel_path, sheet_name, data_list, headers):
    """
    将数据写入Excel，如果文件不存在则创建，否则删除同名sheet后再创建
    """
    # 确保输出目录存在
    output_dir = os.path.dirname(excel_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)  # 自动创建目录
    
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)  # 写入表头
    else:
        wb = openpyxl.load_workbook(excel_path)
        # 删除已存在的同名sheet
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)  # 写入新sheet的表头
    
    # 写入数据行
    for row_data in data_list:
        row = [row_data.get(h, "") for h in headers]
        ws.append(row)
    
    wb.save(excel_path)
    wb.close()   

    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        # 写表头
        ws.append(headers)
    else:
        wb = openpyxl.load_workbook(excel_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        # 写表头
        ws.append(headers)

    for row_data in data_list:
        row = [row_data.get(h, "") for h in headers]
        ws.append(row)

    wb.save(excel_path)
    wb.close()
